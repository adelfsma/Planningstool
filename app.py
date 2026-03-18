import json
import os
import re
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st

# Optional CP-SAT (ortools)
try:
    from ortools.sat.python import cp_model
    ORTOOLS_OK = True
except Exception:
    ORTOOLS_OK = False

APP_VERSION = "19.20"
APP_TITLE = f"Coatinc De Meern - Planning Optimizer v{APP_VERSION}"

st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

# --- Branding ---
_logo_path = os.path.join(os.path.dirname(__file__), "assets", "coatinc_logo.png")
if os.path.exists(_logo_path):
    st.sidebar.image(_logo_path, width=280)

st.markdown(
    """<style>
    div[data-testid='stDataFrame'] div[role='gridcell']{
        white-space: normal !important;
        line-height: 1.15em;
        padding: 0.18rem 0.28rem !important;
        font-size: 0.84rem !important;
    }
    div[data-testid='stDataFrame'] [data-testid='stHeaderCell'] {
        padding: 0.2rem 0.3rem !important;
        font-size: 0.82rem !important;
    }
    div[data-testid='stDataFrame'] {
        max-width: fit-content;
    }
    div[data-testid="stMetric"] {
        background: #f6f8fb;
        border: 1px solid #e6ecf3;
        padding: 12px 14px;
        border-radius: 12px;
    }
    .planner-caption {
        color: #4b5563;
        font-size: 0.95rem;
        margin-top: -0.25rem;
        margin-bottom: 0.75rem;
    }
    .planner-card {
        background: #f8fafc;
        border: 1px solid #e5e7eb;
        border-radius: 12px;
        padding: 14px 16px;
        margin-bottom: 10px;
    }
    .planner-card b {
        color: #0f172a;
    }
    </style>""",
    unsafe_allow_html=True,
)

# ---------- Persistence ----------
APP_DIR = Path(__file__).resolve().parent
SETTINGS_FILE = APP_DIR / "planner_persisted_settings.json"


def easter_sunday(year: int) -> pd.Timestamp:
    # Anonymous Gregorian algorithm
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return pd.Timestamp(year=year, month=month, day=day)


def nl_holidays_for_year(year: int) -> list[pd.Timestamp]:
    easter = easter_sunday(year)
    return sorted(
        [
            pd.Timestamp(year=year, month=1, day=1),   # Nieuwjaar
            easter + timedelta(days=1),                # 2e Paasdag
            pd.Timestamp(year=year, month=4, day=27),  # Koningsdag
            pd.Timestamp(year=year, month=5, day=5),   # Bevrijdingsdag (bewerkbaar/verwijderbaar)
            easter + timedelta(days=39),               # Hemelvaart
            easter + timedelta(days=49),               # 1e Pinksterdag
            easter + timedelta(days=50),               # 2e Pinksterdag
            pd.Timestamp(year=year, month=12, day=25), # 1e Kerstdag
            pd.Timestamp(year=year, month=12, day=26), # 2e Kerstdag
        ]
    )


def default_holiday_text() -> str:
    current_year = pd.Timestamp.now(tz=ZoneInfo("Europe/Amsterdam")).year
    dates = []
    for y in range(current_year - 1, current_year + 3):
        dates.extend(nl_holidays_for_year(y))
    return "\n".join(sorted({d.strftime("%Y-%m-%d") for d in dates}))


def load_persisted_settings() -> dict:
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {"holiday_text": default_holiday_text()}


def save_persisted_settings(data: dict) -> None:
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


persisted = load_persisted_settings()
if "holiday_text" not in st.session_state:
    st.session_state["holiday_text"] = persisted.get("holiday_text", default_holiday_text())

# ---------- Sidebar controls ----------
st.sidebar.header("Basis instellingen")
max_m2 = st.sidebar.number_input("Max m² per dag", value=2000, step=100, min_value=1)
max_kleuren = st.sidebar.number_input("Max kleuren per dag", value=18, step=1, min_value=1)

st.sidebar.subheader("Vrijdagcapaciteit (afwijkend)")
use_friday_override = st.sidebar.checkbox("Gebruik afwijkende capaciteit op vrijdag", value=True)
friday_max_m2 = st.sidebar.number_input("Max m² op vrijdag", value=1600, step=100, min_value=1)
friday_max_kleuren = st.sidebar.number_input("Max kleuren op vrijdag", value=13, step=1, min_value=1)

st.sidebar.subheader("Leverdatumregels")
min_dagen_voor_lever = st.sidebar.number_input(
    "Minimaal # dagen vóór leverdatum (hard)",
    value=2,
    step=1,
    min_value=1,
    help="Hard: productie moet uiterlijk LeverDatum - dit aantal dagen gepland staan.",
)
pref_dagen_voor_lever = st.sidebar.number_input(
    "Voorkeur # dagen vóór leverdatum (soft)",
    value=2,
    step=1,
    min_value=1,
    help="Soft: optimizer probeert deze extra buffer zoveel mogelijk aan te houden.",
)
AANLEVER_PLUS_DAGEN = 1
st.sidebar.caption("Vroegste dag-regel: als Binnen = waar dan direct planbaar; anders vanaf AanleverDatum + 1 dag.")

st.sidebar.subheader("Fixeren")
fixeer_morgen = st.sidebar.checkbox(
    "Fixeer ook volgende werkdag",
    value=True,
    help="Vandaag en eerder zijn altijd gefixeerd. Met dit vinkje wordt de volgende planbare werkdag ook op slot gezet.",
)
fixeer_overmorgen = st.sidebar.checkbox(
    "Fixeer ook 2e volgende werkdag",
    value=False,
    help="Zet ook de 2e volgende planbare werkdag op slot (geen herplanning / geen nieuwe orders op die dag).",
)

st.sidebar.subheader("Kalender")
exclude_weekends = st.sidebar.checkbox("Weekenden uitsluiten", value=True)
exclude_holidays = st.sidebar.checkbox("Feestdagen / sluitingsdagen uitsluiten", value=True)
st.sidebar.caption(
    "Onderstaand vak is standaard gevuld met NL-feestdagen. Je kunt regels verwijderen of toevoegen. "
    "Klik op Opslaan om jouw lijst blijvend te bewaren."
)
holiday_text = st.sidebar.text_area(
    "Feestdagen en sluitingsdagen (YYYY-MM-DD, één per regel)",
    key="holiday_text",
    height=240,
)
_save_col, _reset_col = st.sidebar.columns(2)
if _save_col.button("Opslaan lijst", use_container_width=True):
    save_persisted_settings({"holiday_text": st.session_state["holiday_text"]})
    st.sidebar.success("Feestdagen/sluitingsdagen opgeslagen.")
if _reset_col.button("Reset NL", use_container_width=True):
    st.session_state["holiday_text"] = default_holiday_text()
    save_persisted_settings({"holiday_text": st.session_state["holiday_text"]})
    st.rerun()

st.sidebar.subheader("Optimizer")
optimizer = st.sidebar.selectbox("Methode", ["Heuristiek (snel)", "CP-SAT (beste oplossing)"], index=1)
time_limit = st.sidebar.slider(
    "CP-SAT tijdslimiet (sec)",
    5,
    120,
    20,
    step=5,
    disabled=(optimizer != "CP-SAT (beste oplossing)" or not ORTOOLS_OK),
)

uploaded = st.file_uploader("Upload planning Excel", type=["xlsx"])

# ---------- Helpers ----------
def parse_holidays(txt: str) -> set[pd.Timestamp]:
    dates = set()
    for line in txt.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            dates.add(pd.to_datetime(line).normalize())
        except Exception:
            pass
    return dates


def drop_deelorder_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None:
        return df_in
    cols = [c for c in df_in.columns if "deelorder" not in str(c).lower()]
    return df_in[cols]


def fmt_m2(x):
    try:
        return f"{float(x):.1f}".replace(".", ",")
    except Exception:
        return x


def coerce_m2_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    df2 = df_in.copy()
    for c in df2.columns:
        low = str(c).lower()
        if "m2" in low or "m²" in low:
            df2[c] = pd.to_numeric(df2[c], errors="coerce").round(1)
    return df2


def fmt_df_m2(df, cols):
    df2 = df.copy()
    for c in cols:
        if c in df2.columns:
            df2[c] = pd.to_numeric(df2[c], errors="coerce").round(1).apply(fmt_m2)
    return df2


def format_pct(x):
    try:
        return f"{float(x) * 100:.0f}%"
    except Exception:
        return x


def normalize_dates(df_in: pd.DataFrame) -> pd.DataFrame:
    for c in df_in.columns:
        low = str(c).lower()
        if ("datum" in low) or low.endswith("dag"):
            try:
                if pd.api.types.is_datetime64_any_dtype(df_in[c]):
                    df_in[c] = pd.to_datetime(df_in[c], errors="coerce").dt.normalize()
            except Exception:
                pass
    return df_in


def format_dates_ddmmyyyy(df_in: pd.DataFrame) -> pd.DataFrame:
    df2 = df_in.copy()
    for c in df2.columns:
        low = str(c).lower()
        if ("datum" in low) or low.endswith("dag"):
            try:
                if pd.api.types.is_datetime64_any_dtype(df2[c]):
                    df2[c] = pd.to_datetime(df2[c], errors="coerce").dt.strftime("%d-%m-%Y")
            except Exception:
                pass
    return df2


def make_column_config(df: pd.DataFrame):
    cfg = {}
    medium_cols = {
        "kleur",
        "stoplicht m2",
        "stoplicht kleuren",
        "belasting m2 pct",
        "belasting kleur pct",
        "referentieklant",
        "orderids",
        "actieblok",
        "adviestoelichting",
    }
    large_cols = {"referentieklant"}
    date_cols = {
        "oudeplandatum",
        "nieuweplandatum",
        "leverdatum",
        "aanleverdatum",
        "vroegste dag",
        "laatste toegestane dag",
        "vroegstedag",
        "laatstetoegestanedag",
        "dag",
    }
    for col in df.columns:
        low = str(col).lower()
        if low in large_cols:
            cfg[col] = st.column_config.TextColumn(col, width="large")
        elif low in medium_cols or low in date_cols:
            cfg[col] = st.column_config.TextColumn(col, width="medium")
        else:
            cfg[col] = st.column_config.TextColumn(col, width="small")
    return cfg


def render_df(df_in: pd.DataFrame, *, height: int | str = "content"):
    if height is None:
        height = "content"
    st.dataframe(
        format_dates_ddmmyyyy(df_in),
        use_container_width=False,
        hide_index=True,
        height=height,
        column_config=make_column_config(df_in),
    )


def compact_dagoverzicht_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    df_out = df_in.copy()
    keep_map = [
        ("Dag", "Dag"),
        ("Orders", "aantal orders"),
        ("m2", "m2"),
        ("Kleuren", "kleuren"),
        ("Stoplicht_m2", "Stoplicht m2"),
        ("Stoplicht_kleuren", "stoplicht kleuren"),
        ("Belasting_m2_pct", "Belasting m2 pct"),
        ("Belasting_kleur_pct", "belasting kleur pct"),
    ]
    cols = []
    for src, dst in keep_map:
        if src in df_out.columns:
            cols.append(src)
    df_out = df_out[cols]
    rename_map = {src: dst for src, dst in keep_map if src in df_out.columns}
    return df_out.rename(columns=rename_map)


def compact_detailplanning_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    df_out = df_in.copy()
    keep_map = [
        ("Naam", "Naam"),
        ("OrderID", "OrderID"),
        ("VolgNr", "Volgnummer"),
        ("Kleur", "Kleur"),
        ("M2", "M2"),
        ("LeverDatum", "Leverdatum"),
        ("AanleverDatum", "Aanleverdatum"),
        ("OudePlanDatum", "oudePlanDatum"),
        ("NieuwePlanDatum", "NieuwePlanDatum"),
        ("Binnen", "Binnen"),
        ("ReferentieKlant", "ReferentieKlant"),
        ("LaatsteToegestaneDag", "Laatste toegestane dag"),
        ("VroegsteDag", "vroegste dag"),
        ("Gewijzigd", "gewijzigd"),
        ("Gefixeerd", "gefixeerd"),
    ]
    cols = [src for src, _ in keep_map if src in df_out.columns]
    df_out = df_out[cols]
    rename_map = {src: dst for src, dst in keep_map if src in df_out.columns}
    return df_out.rename(columns=rename_map)


def is_weekend(d: pd.Timestamp) -> bool:
    return d.weekday() >= 5


def is_counted_workday(d: pd.Timestamp, holidays_set: set[pd.Timestamp]) -> bool:
    dn = pd.Timestamp(d).normalize()
    return dn.weekday() < 5 and dn not in holidays_set


def subtract_workdays(d: pd.Timestamp, n: int, holidays_set: set[pd.Timestamp]) -> pd.Timestamp:
    if pd.isna(d):
        return pd.NaT
    cur = pd.Timestamp(d).normalize()
    count = 0
    while count < int(n):
        cur = cur - timedelta(days=1)
        if is_counted_workday(cur, holidays_set):
            count += 1
    return cur


def add_workdays(d: pd.Timestamp, n: int, holidays_set: set[pd.Timestamp]) -> pd.Timestamp:
    if pd.isna(d):
        return pd.NaT
    cur = pd.Timestamp(d).normalize()
    count = 0
    while count < int(n):
        cur = cur + timedelta(days=1)
        if is_counted_workday(cur, holidays_set):
            count += 1
    return cur


def add_calendar_days(d: pd.Timestamp, n: int) -> pd.Timestamp:
    if pd.isna(d):
        return pd.NaT
    return pd.Timestamp(d).normalize() + timedelta(days=int(n))


def make_working_days(start, end, excl_weekends, holidays_set, excl_holidays):
    days = pd.date_range(start=start, end=end, freq="D")
    out = []
    for d in days:
        dn = d.normalize()
        if excl_weekends and is_weekend(dn):
            continue
        if excl_holidays and dn in holidays_set:
            continue
        out.append(dn)
    return out


def next_planning_days(base_day: pd.Timestamp, n: int, excl_weekends: bool, holidays_set: set[pd.Timestamp], excl_holidays: bool):
    out = []
    cur = pd.Timestamp(base_day).normalize()
    while len(out) < n:
        cur = cur + timedelta(days=1)
        if excl_weekends and is_weekend(cur):
            continue
        if excl_holidays and cur in holidays_set:
            continue
        out.append(cur)
    return out


def get_day_caps(d: pd.Timestamp):
    if use_friday_override and pd.Timestamp(d).weekday() == 4:
        return int(friday_max_m2), int(friday_max_kleuren)
    return int(max_m2), int(max_kleuren)


def safe_bool(x):
    if isinstance(x, bool):
        return x
    if pd.isna(x):
        return False
    s = str(x).strip().lower()
    if s in ("true", "waar", "1", "yes", "y", "ja"):
        return True
    if s in ("false", "onwaar", "0", "no", "n", "nee"):
        return False
    return False


# ---------- Main ----------
if not uploaded:
    st.info(
        "Upload een Excel met minimaal kolommen: PlanDatum, LeverDatum, Kleur, M2, Binnen. "
        "Nieuwe regel: als Binnen = onwaar mag de plandatum niet eerder liggen dan 5 werkdagen vóór de leverdatum (instelbaar)."
    )
    st.stop()

df = pd.read_excel(uploaded)
df = coerce_m2_columns(df)

required = ["PlanDatum", "LeverDatum", "Kleur", "M2", "Binnen"]
missing = [c for c in required if c not in df.columns]
if missing:
    st.error(f"Kolommen ontbreken: {', '.join(missing)}")
    st.stop()

if "OrderID" not in df.columns:
    df["OrderID"] = range(1, len(df) + 1)

# Parse dates robustly / timezone-safe
for col in ["PlanDatum", "LeverDatum", "AanleverDatum"]:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")
        try:
            if getattr(df[col].dt, "tz", None) is not None:
                df[col] = df[col].dt.tz_convert("Europe/Amsterdam").dt.tz_localize(None)
        except Exception:
            try:
                df[col] = pd.to_datetime(df[col].astype(str), errors="coerce")
            except Exception:
                pass

# Normalize Binnen
df["_BinnenBool"] = df["Binnen"].apply(safe_bool)

vandaag = pd.Timestamp.now(tz=ZoneInfo("Europe/Amsterdam")).normalize().tz_localize(None)

holidays = parse_holidays(st.session_state["holiday_text"])
volgende_werkdagen = next_planning_days(vandaag, 2, exclude_weekends, holidays, exclude_holidays)
volgende_werkdag = volgende_werkdagen[0] if len(volgende_werkdagen) >= 1 else vandaag + timedelta(days=1)
tweede_volgende_werkdag = volgende_werkdagen[1] if len(volgende_werkdagen) >= 2 else volgende_werkdag + timedelta(days=1)

# Fix rules
plan_norm = pd.to_datetime(df["PlanDatum"], errors="coerce").dt.normalize()
df["Gefixeerd"] = plan_norm.le(vandaag).fillna(False)
if fixeer_morgen:
    df.loc[plan_norm == volgende_werkdag, "Gefixeerd"] = True
if fixeer_overmorgen:
    df.loc[plan_norm == tweede_volgende_werkdag, "Gefixeerd"] = True

locked_days = set(pd.to_datetime(df.loc[df["Gefixeerd"], "PlanDatum"], errors="coerce").dt.normalize().dropna().unique())

max_lever = df["LeverDatum"].max()
if pd.isna(max_lever):
    st.error("LeverDatum bevat lege/ongeldige waarden.")
    st.stop()

horizon_end = (max_lever + timedelta(days=14)).normalize()
start = (vandaag + timedelta(days=1)).normalize()

working_days = make_working_days(start, horizon_end, exclude_weekends, holidays, exclude_holidays)
candidate_days = [d for d in working_days if d not in locked_days]

if len(candidate_days) == 0:
    st.error("Geen planbare dagen in de horizon (mogelijk alles geblokkeerd door weekends/feestdagen/fixaties).")
    st.stop()

# Precompute fixed day usage
fixed = df[df["Gefixeerd"]].copy()
fixed["PlanDatumN"] = pd.to_datetime(fixed["PlanDatum"], errors="coerce").dt.normalize()
fixed_m2 = fixed.groupby("PlanDatumN")["M2"].sum().to_dict()
fixed_colors = fixed.groupby("PlanDatumN")["Kleur"].nunique().to_dict()

# Day warning for fixed overload
overload_days = []
for d in sorted(locked_days):
    m2d = float(fixed_m2.get(d, 0.0))
    cd = int(fixed_colors.get(d, 0))
    cap_m2_d, cap_k_d = get_day_caps(pd.Timestamp(d))
    if m2d > cap_m2_d or cd > cap_k_d:
        overload_days.append((d, m2d, cd, cap_m2_d, cap_k_d))

# Compute deadlines

df["LaatsteToegestaneDag"] = df["LeverDatum"].dt.normalize() - pd.to_timedelta(int(min_dagen_voor_lever), unit="D")
df["VoorkeurLaatsteDag"] = df["LeverDatum"].dt.normalize() - pd.to_timedelta(int(pref_dagen_voor_lever), unit="D")


def earliest_day(row):
    e = start
    if bool(row.get("_BinnenBool", False)):
        return e
    aanlever = pd.to_datetime(row.get("AanleverDatum"), errors="coerce")
    if pd.notna(aanlever):
        e = max(e, add_calendar_days(aanlever, AANLEVER_PLUS_DAGEN))
    return e


def feasible_days_for_row(row, days, ignore_latest=False):
    earliest = row.get("VroegsteDag", pd.NaT)
    latest = row.get("LaatsteToegestaneDag", pd.NaT)
    if pd.isna(earliest):
        earliest = start
    if ignore_latest:
        return [d for d in days if d >= earliest]
    if pd.isna(latest):
        return []
    return [d for d in days if d >= earliest and d <= latest]


df["VroegsteDag"] = df.apply(earliest_day, axis=1)
original_plan = pd.to_datetime(df["PlanDatum"], errors="coerce").dt.normalize().copy()

df["Regel_AanleverdatumVanToepassing"] = (~df["_BinnenBool"]).map(lambda x: "Ja" if bool(x) else "Nee")

# ---------- Heuristic ----------
def solve_heuristic(df_in: pd.DataFrame):
    dfh = df_in.copy()
    day_m2 = {d: float(fixed_m2.get(d, 0.0)) for d in candidate_days}
    day_colors = {d: set(fixed.loc[fixed["PlanDatumN"] == d, "Kleur"].astype(str).tolist()) for d in candidate_days}
    fixed_candidate = fixed[fixed["PlanDatumN"].isin(candidate_days)].copy()
    day_order_ids = {d: set(fixed_candidate.loc[fixed_candidate["PlanDatumN"] == d, "OrderID"].astype(str).tolist()) for d in candidate_days}
    day_order_color = {
        d: set(
            zip(
                fixed_candidate.loc[fixed_candidate["PlanDatumN"] == d, "OrderID"].astype(str),
                fixed_candidate.loc[fixed_candidate["PlanDatumN"] == d, "Kleur"].astype(str),
            )
        )
        for d in candidate_days
    }

    new_dates = {}
    unplanned_reason = {}

    dfh2 = dfh.sort_values(["LaatsteToegestaneDag", "LeverDatum", "M2"], ascending=[True, True, False])

    for idx, row in dfh2.iterrows():
        if row["Gefixeerd"] or pd.isna(row["PlanDatum"]):
            new_dates[idx] = row["PlanDatum"].normalize() if pd.notna(row["PlanDatum"]) else pd.NaT
            continue

        latest = row["LaatsteToegestaneDag"]
        earliest = row["VroegsteDag"]
        pref_latest = row["VoorkeurLaatsteDag"]

        feas_days = feasible_days_for_row(row, candidate_days, ignore_latest=False)

        if not feas_days:
            new_dates[idx] = pd.NaT
            if pd.notna(earliest) and pd.notna(latest) and earliest > latest:
                unplanned_reason[idx] = "Venster ongeldig: vroegste dag ligt na hard deadline"
            else:
                unplanned_reason[idx] = "Niet planbaar binnen kalender/deadline"
            continue

        best = None
        for d in feas_days:
            m2 = float(row["M2"])
            cap_m2_d, cap_k_d = get_day_caps(pd.Timestamp(d))
            if day_m2[d] + m2 > cap_m2_d:
                continue
            colors = day_colors[d]
            kleur = str(row["Kleur"])
            order_id = str(row.get("OrderID", ""))
            order_color = (order_id, kleur)
            new_color = kleur not in colors
            if (len(colors) + (1 if new_color else 0)) > cap_k_d:
                continue

            after_pref = 1 if pd.notna(pref_latest) and d > pref_latest else 0
            add_color = 1 if new_color else 0
            new_order_color_day = 0 if order_color in day_order_color[d] else 1
            new_order_day = 0 if order_id in day_order_ids[d] else 1
            remcap = cap_m2_d - (day_m2[d] + m2)
            score = (after_pref, add_color, new_order_color_day, new_order_day, remcap)

            if best is None or score < best[0]:
                best = (score, d)

        if best is None:
            new_dates[idx] = pd.NaT
            unplanned_reason[idx] = "Capaciteit/kleur-limiet"
            continue

        d = best[1]
        new_dates[idx] = d
        day_m2[d] += float(row["M2"])
        day_colors[d].add(str(row["Kleur"]))
        order_id = str(row.get("OrderID", ""))
        kleur = str(row["Kleur"])
        day_order_ids[d].add(order_id)
        day_order_color[d].add((order_id, kleur))

    return new_dates, unplanned_reason


# ---------- CP-SAT ----------
def solve_cpsat(df_in: pd.DataFrame):
    if not ORTOOLS_OK:
        return None, {i: "OR-Tools niet beschikbaar" for i in df_in.index}

    dfi = df_in.copy()
    var_idx = [i for i, r in dfi.iterrows() if not r["Gefixeerd"]]
    days = candidate_days[:]
    T = len(days)

    fixed_m2_arr = [float(fixed_m2.get(d, 0.0)) for d in days]
    fixed_color_sets = [set(fixed.loc[fixed["PlanDatumN"] == d, "Kleur"].astype(str).tolist()) for d in days]

    colors = sorted(dfi["Kleur"].dropna().astype(str).unique().tolist())
    color_to_k = {c: k for k, c in enumerate(colors)}
    K = len(colors)

    model = cp_model.CpModel()

    x = {}
    unassigned = {}
    feasible_days_per_order = {}
    for i in var_idx:
        unassigned[i] = model.NewBoolVar(f"unassigned_{i}")
        earliest = dfi.at[i, "VroegsteDag"]
        latest = dfi.at[i, "LaatsteToegestaneDag"]
        feasible_days = feasible_days_for_row(dfi.loc[i], days, ignore_latest=False)
        feasible_ts = [t for t, d in enumerate(days) if d in feasible_days]
        feasible_days_per_order[i] = feasible_ts
        for t in feasible_ts:
            x[(i, t)] = model.NewBoolVar(f"x_{i}_{t}")
        model.Add(sum(x[(i, t)] for t in feasible_ts) + unassigned[i] == 1)

    for t, d in enumerate(days):
        m2_terms = []
        for i in var_idx:
            if (i, t) not in x:
                continue
            m2 = int(round(float(dfi.at[i, "M2"])))
            m2_terms.append(m2 * x[(i, t)])
        cap_m2_t, _ = get_day_caps(d)
        model.Add(sum(m2_terms) + int(round(fixed_m2_arr[t])) <= int(cap_m2_t))

    y = {}
    for k, c in enumerate(colors):
        for t, d in enumerate(days):
            y[(k, t)] = model.NewBoolVar(f"y_{k}_{t}")
            if c in fixed_color_sets[t]:
                model.Add(y[(k, t)] == 1)

    for i in var_idx:
        c = str(dfi.at[i, "Kleur"])
        k = color_to_k.get(c, None)
        if k is None:
            continue
        for t in feasible_days_per_order[i]:
            model.Add(y[(k, t)] >= x[(i, t)])

    for t in range(T):
        _, cap_k_t = get_day_caps(days[t])
        model.Add(sum(y[(k, t)] for k in range(K)) <= int(cap_k_t))

    order_labels = sorted(set(dfi["OrderID"].astype(str).tolist()))
    z_order = {}
    for o in order_labels:
        fixed_days_o = set()
        fixed_rows = dfi[(dfi["Gefixeerd"]) & (dfi["OrderID"].astype(str) == o)]
        for _, fr in fixed_rows.iterrows():
            pdn = pd.to_datetime(fr["PlanDatum"], errors="coerce")
            if pd.notna(pdn):
                fixed_days_o.add(pdn.normalize())
        safe_o = re.sub(r"[^A-Za-z0-9_]+", "_", o)[:40]
        for t, d in enumerate(days):
            z_order[(o, t)] = model.NewBoolVar(f"z_order_{safe_o}_{t}")
            if d in fixed_days_o:
                model.Add(z_order[(o, t)] == 1)
    for i in var_idx:
        o = str(dfi.at[i, "OrderID"])
        for t in feasible_days_per_order[i]:
            model.Add(z_order[(o, t)] >= x[(i, t)])

    order_color_labels = sorted(set((str(r["OrderID"]), str(r["Kleur"])) for _, r in dfi.iterrows()))
    z_order_color = {}
    for o, c in order_color_labels:
        fixed_days_oc = set()
        fixed_rows = dfi[(dfi["Gefixeerd"]) & (dfi["OrderID"].astype(str) == o) & (dfi["Kleur"].astype(str) == c)]
        for _, fr in fixed_rows.iterrows():
            pdn = pd.to_datetime(fr["PlanDatum"], errors="coerce")
            if pd.notna(pdn):
                fixed_days_oc.add(pdn.normalize())
        safe_o = re.sub(r"[^A-Za-z0-9_]+", "_", o)[:30]
        safe_c = re.sub(r"[^A-Za-z0-9_]+", "_", c)[:30]
        for t, d in enumerate(days):
            z_order_color[((o, c), t)] = model.NewBoolVar(f"z_oc_{safe_o}_{safe_c}_{t}")
            if d in fixed_days_oc:
                model.Add(z_order_color[((o, c), t)] == 1)
    for i in var_idx:
        o = str(dfi.at[i, "OrderID"])
        c = str(dfi.at[i, "Kleur"])
        for t in feasible_days_per_order[i]:
            model.Add(z_order_color[((o, c), t)] >= x[(i, t)])

    BIG = 1_000_000
    W_COLOR = 1000
    W_ORDER_COLOR_DAY = 20
    W_ORDER_DAY = 8
    W_PREF = 1
    obj_terms = [BIG * unassigned[i] for i in var_idx]
    obj_terms += [W_COLOR * y[(k, t)] for k in range(K) for t in range(T)]
    obj_terms += [W_ORDER_COLOR_DAY * z_order_color[(oc, t)] for oc in order_color_labels for t in range(T)]
    obj_terms += [W_ORDER_DAY * z_order[(o, t)] for o in order_labels for t in range(T)]

    for i in var_idx:
        pref_latest = dfi.at[i, "VoorkeurLaatsteDag"]
        for t in feasible_days_per_order[i]:
            d = days[t]
            if pd.notna(pref_latest) and d > pref_latest:
                penalty = int((d - pref_latest).days)
                obj_terms.append(W_PREF * penalty * x[(i, t)])

    model.Minimize(sum(obj_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit)
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    new_dates = {}
    reasons = {}

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for i in var_idx:
            new_dates[i] = pd.NaT
            reasons[i] = "Geen oplossing gevonden (tijdslimiet/constraints)"
        for i, r in dfi.iterrows():
            if r["Gefixeerd"]:
                new_dates[i] = r["PlanDatum"].normalize() if pd.notna(r["PlanDatum"]) else pd.NaT
        return new_dates, reasons

    for i, r in dfi.iterrows():
        if r["Gefixeerd"]:
            new_dates[i] = r["PlanDatum"].normalize() if pd.notna(r["PlanDatum"]) else pd.NaT

    for i in var_idx:
        if solver.Value(unassigned[i]) == 1:
            new_dates[i] = pd.NaT
            earliest = dfi.at[i, "VroegsteDag"]
            latest = dfi.at[i, "LaatsteToegestaneDag"]
            if pd.notna(earliest) and pd.notna(latest) and earliest > latest:
                reasons[i] = "Venster ongeldig: vroegste dag ligt na hard deadline"
            else:
                reasons[i] = "Niet planbaar binnen constraints"
            continue
        assigned_t = None
        for t in feasible_days_per_order[i]:
            if solver.Value(x[(i, t)]) == 1:
                assigned_t = t
                break
        new_dates[i] = days[assigned_t] if assigned_t is not None else pd.NaT

    return new_dates, reasons


# ---------- Run solve ----------
if optimizer.startswith("CP-SAT"):
    if not ORTOOLS_OK:
        st.error("CP-SAT gekozen, maar OR-Tools (ortools) is niet beschikbaar in deze Python omgeving. Kies 'Heuristiek (snel)' of installeer ortools.")
        st.stop()
    new_dates, reasons = solve_cpsat(df)
else:
    new_dates, reasons = solve_heuristic(df)

df["NieuwePlanDatum"] = pd.Series(new_dates)
df["OudePlanDatum"] = original_plan
df["Gewijzigd"] = (df["NieuwePlanDatum"] != df["OudePlanDatum"]) & df["NieuwePlanDatum"].notna()

df["Gefixeerd_JaNee"] = df["Gefixeerd"].map(lambda x: "Ja" if bool(x) else "Nee")

df["HardDeadlineOK"] = df["NieuwePlanDatum"].notna() & (df["NieuwePlanDatum"] <= df["LaatsteToegestaneDag"])
late = df[(df["NieuwePlanDatum"].isna()) | (~df["HardDeadlineOK"])].copy()
if len(late):
    late["Reden"] = late.index.map(lambda i: reasons.get(i, "Hard deadline overschreden"))
else:
    late["Reden"] = ""



LATE_REMOVE_COLS = [
    "DeelorderID", "DeelOrderDienstID", "ProduktieLijn", "LeverUur", "IngeplandDoorID",
    "StatusDiensten", "StatusDienstenGereed", "Omschrijving", "PoederID", "Brightness",
    "MaxDikte", "LeverDatumTypeID", "MateriaalSoort", "AantalLagenPoedercoating", "Uren",
    "ProduktielijnID", "Trav", "Poed", "Omzet", "PlanstatusDefinitief", "M2_Def",
    "Uren_Def", "Trav_Def", "Poed_Def", "Omzet_Def", "Kleurbitmap", "Gereed", "Colli",
    "Vak", "Lgn", "Grond", "Stralen", "StralenGereed", "Poetsen", "PoetsenGereed",
    "LeverDatumTypeAfkorting", "VoorkeurLaatsteDag", "VroegsteDag",
    "Regel_AanleverdatumVanToepassing", "HardDeadlineOK", "OrderTypeAfkorting"
]

def compact_late_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    if df_in is None or len(df_in.columns) == 0:
        return df_in
    df_out = df_in.copy()
    keep_cols = [c for c in df_out.columns if c not in LATE_REMOVE_COLS]
    df_out = df_out[keep_cols]

    ordered_front = [
        "OrderID", "Naam", "ReferentieKlant", "Kleur", "M2", "LeverDatum", "Reden",
        "OudePlanDatum", "AdviesPlanDatum", "AdviesLeverDatum", "AdviesToelichting", "Ralzoekcode",
        "WerkdagenTeLaat", "AdviesActie"
    ]
    front = [c for c in ordered_front if c in df_out.columns]
    rest = [c for c in df_out.columns if c not in front]
    return df_out[front + rest]

def build_late_advice(df_all: pd.DataFrame, late_df: pd.DataFrame) -> pd.DataFrame:
    late_out = late_df.copy()
    if len(late_out) == 0:
        return late_out

    advice_end = (df_all["LeverDatum"].max() + timedelta(days=90)).normalize()
    advice_days = make_working_days(start, advice_end, exclude_weekends, holidays, exclude_holidays)
    advice_days = [d for d in advice_days if d not in locked_days]

    planned = df_all[df_all["EffectivePlanDag"].notna()].copy() if "EffectivePlanDag" in df_all.columns else pd.DataFrame()
    occ_m2 = planned.groupby("EffectivePlanDag")["M2"].sum().to_dict() if len(planned) else {}
    occ_colors = planned.groupby("EffectivePlanDag")["Kleur"].apply(lambda s: set(s.astype(str))).to_dict() if len(planned) else {}

    advies_plan = {}
    advies_lever = {}
    werkdagen_te_laat = {}
    advies_actie = {}
    advies_toelichting = {}

    late_sorted = late_out.sort_values([c for c in ["LeverDatum", "M2"] if c in late_out.columns], ascending=[True, False])

    for idx, row in late_sorted.iterrows():
        earliest = row.get("VroegsteDag", start)
        if pd.isna(earliest):
            earliest = start
        kleur = str(row.get("Kleur", ""))
        m2 = float(row.get("M2", 0.0) or 0.0)

        advice_feasible_days = feasible_days_for_row(row, advice_days, ignore_latest=True)

        found = pd.NaT
        for d in advice_feasible_days:
            cap_m2_d, cap_k_d = get_day_caps(pd.Timestamp(d))
            used_m2 = float(occ_m2.get(d, 0.0))
            used_colors = set(occ_colors.get(d, set()))
            add_color = 0 if kleur in used_colors else 1
            if used_m2 + m2 <= cap_m2_d and len(used_colors) + add_color <= cap_k_d:
                found = d
                occ_m2[d] = used_m2 + m2
                used_colors.add(kleur)
                occ_colors[d] = used_colors
                break

        advies_plan[idx] = found
        if pd.notna(found):
            lever = add_workdays(found, int(min_dagen_voor_lever), holidays)
            advies_lever[idx] = lever
            req = pd.to_datetime(row.get("LeverDatum"), errors="coerce")
            if pd.notna(req) and pd.notna(lever):
                werkdagen_te_laat[idx] = sum(1 for d in pd.date_range(req.normalize() + timedelta(days=1), lever.normalize(), freq="D") if is_counted_workday(d, holidays))
            else:
                werkdagen_te_laat[idx] = pd.NA
            if pd.notna(werkdagen_te_laat[idx]) and werkdagen_te_laat[idx] <= 2:
                advies_actie[idx] = "Klantdatum herbevestigen"
            elif pd.notna(werkdagen_te_laat[idx]) and werkdagen_te_laat[idx] <= 5:
                advies_actie[idx] = "Verplaatsen naar eerstvolgende haalbare dag"
            else:
                advies_actie[idx] = "Nieuwe leverbelofte afstemmen"
            req_str = req.strftime('%d-%m-%Y') if pd.notna(req) else 'onbekend'
            advies_toelichting[idx] = (
                f"Gevraagde leverdatum: {req_str}. "
                f"Plan op {found.strftime('%d-%m-%Y')}; uitlevering haalbaar op {lever.strftime('%d-%m-%Y')}."
            )
        else:
            advies_lever[idx] = pd.NaT
            werkdagen_te_laat[idx] = pd.NA
            advies_actie[idx] = "Handmatige beoordeling nodig"
            req = pd.to_datetime(row.get("LeverDatum"), errors="coerce")
            req_str = req.strftime('%d-%m-%Y') if pd.notna(req) else 'onbekend'
            advies_toelichting[idx] = (
                f"Gevraagde leverdatum: {req_str}. "
                "Geen haalbare dag gevonden binnen de advieshorizon van 90 dagen."
            )

    late_out["AdviesPlanDatum"] = late_out.index.map(advies_plan.get)
    late_out["AdviesLeverDatum"] = late_out.index.map(advies_lever.get)
    late_out["WerkdagenTeLaat"] = late_out.index.map(werkdagen_te_laat.get)
    late_out["AdviesActie"] = late_out.index.map(advies_actie.get)
    late_out["AdviesToelichting"] = late_out.index.map(advies_toelichting.get)
    return late_out

herpland = df[df["Gewijzigd"]].copy()
herpland["Verschil_dagen"] = (herpland["NieuwePlanDatum"] - herpland["OudePlanDatum"]).dt.days
herpland["Verplaatsrichting"] = herpland["Verschil_dagen"].apply(lambda x: "Later" if pd.notna(x) and x > 0 else "Eerder")
herpland["Verplaatsgroep"] = (
    herpland["OudePlanDatum"].dt.strftime("%d-%m-%Y")
    + " → "
    + herpland["NieuwePlanDatum"].dt.strftime("%d-%m-%Y")
)
herpland["Actie"] = "Verplaats order naar " + herpland["NieuwePlanDatum"].dt.strftime("%d-%m-%Y")

# ---------- Effective planning / day summaries ----------
df["EffectivePlanDag"] = df.apply(
    lambda r: (r["PlanDatum"].normalize() if bool(r["Gefixeerd"]) and pd.notna(r["PlanDatum"]) else r["NieuwePlanDatum"]),
    axis=1,
)

late = build_late_advice(df, late)

day_base = df[df["EffectivePlanDag"].notna()].copy()
day_base["Dag"] = day_base["EffectivePlanDag"].dt.normalize()
fixed_part = day_base[day_base["Gefixeerd"].astype(bool)].copy()
new_part = day_base[~day_base["Gefixeerd"].astype(bool)].copy()

d_fix = fixed_part.groupby("Dag").agg(
    m2_gefixeerd=("M2", "sum"),
    kleuren_gefixeerd=("Kleur", pd.Series.nunique),
    orders_gefixeerd=("OrderID", "count"),
)

d_new = new_part.groupby("Dag").agg(
    m2_nieuw=("M2", "sum"),
    kleuren_nieuw=("Kleur", pd.Series.nunique),
    orders_nieuw=("OrderID", "count"),
)

dagsamenvatting = pd.concat([d_fix, d_new], axis=1).fillna(0).reset_index()
dagsamenvatting["Orders"] = dagsamenvatting["orders_gefixeerd"].astype(int) + dagsamenvatting["orders_nieuw"].astype(int)
dagsamenvatting["m2"] = dagsamenvatting["m2_gefixeerd"].astype(float) + dagsamenvatting["m2_nieuw"].astype(float)
d_tot_colors = day_base.groupby("Dag")["Kleur"].nunique().rename("Kleuren").reset_index()
dagsamenvatting = dagsamenvatting.merge(d_tot_colors, on="Dag", how="left")
dagsamenvatting["Cap_m2"] = dagsamenvatting["Dag"].apply(lambda d: float(get_day_caps(pd.Timestamp(d))[0]))
dagsamenvatting["Cap_kleuren"] = dagsamenvatting["Dag"].apply(lambda d: int(get_day_caps(pd.Timestamp(d))[1]))
dagsamenvatting["Belasting_m2_pct"] = (dagsamenvatting["m2"] / dagsamenvatting["Cap_m2"]).fillna(0)
dagsamenvatting["Belasting_kleur_pct"] = (dagsamenvatting["Kleuren"] / dagsamenvatting["Cap_kleuren"]).fillna(0)

def stoplicht(pct: float) -> str:
    if pct > 1:
        return "🔴"
    if pct >= 0.9:
        return "🟧"
    return "🟩"


dagsamenvatting["Stoplicht_m2"] = dagsamenvatting["Belasting_m2_pct"].apply(stoplicht)
dagsamenvatting["Stoplicht_kleuren"] = dagsamenvatting["Belasting_kleur_pct"].apply(stoplicht)
dagsamenvatting = dagsamenvatting.sort_values("Dag")

kleurblokken = (
    day_base.groupby(["Dag", "Kleur"]).agg(
        Orders=("OrderID", "count"),
        m2=("M2", "sum"),
        Gefixeerd=("Gefixeerd_JaNee", lambda s: "Ja" if (s == "Ja").any() else "Nee"),
    )
    .reset_index()
    .sort_values(["Dag", "m2"], ascending=[True, False])
)

kleurbelasting_week = pd.DataFrame()
if len(day_base):
    week_base = day_base.copy()
    iso = week_base["Dag"].dt.isocalendar()
    week_base["Jaar"] = iso.year.astype(int)
    week_base["WeekNr"] = iso.week.astype(int)
    kleurbelasting_week = (
        week_base.groupby(["Jaar", "WeekNr", "Kleur"], dropna=False)
        .agg(
            Orders=("OrderID", "count"),
            m2=("M2", "sum"),
            DagenGepland=("Dag", pd.Series.nunique),
            EersteDag=("Dag", "min"),
            LaatsteDag=("Dag", "max"),
        )
        .reset_index()
        .sort_values(["Jaar", "WeekNr", "m2", "Kleur"], ascending=[True, True, False, True])
    )
    kleurbelasting_week["Week"] = kleurbelasting_week["Jaar"].astype(str) + "-W" + kleurbelasting_week["WeekNr"].astype(str).str.zfill(2)

# ---------- Planner-friendly replan views ----------
verplaatsblokken = pd.DataFrame()
verplaatsblokken_kleur = pd.DataFrame()
if len(herpland):
    verplaatsblokken = (
        herpland.groupby(["OudePlanDatum", "NieuwePlanDatum", "Verplaatsrichting"], dropna=False)
        .agg(
            Orders=("OrderID", "count"),
            Kleuren=("Kleur", pd.Series.nunique),
            m2=("M2", "sum"),
            OrderIDs=("OrderID", lambda s: ", ".join(map(str, pd.Series(s).astype(str).tolist()))),
        )
        .reset_index()
        .sort_values(["OudePlanDatum", "NieuwePlanDatum"])
    )
    verplaatsblokken["Actieblok"] = (
        "Verplaats " + verplaatsblokken["Orders"].astype(int).astype(str)
        + " order(s) van " + verplaatsblokken["OudePlanDatum"].dt.strftime("%d-%m-%Y")
        + " naar " + verplaatsblokken["NieuwePlanDatum"].dt.strftime("%d-%m-%Y")
    )

    verplaatsblokken_kleur = (
        herpland.groupby(["Kleur", "OudePlanDatum", "NieuwePlanDatum", "Verplaatsrichting"], dropna=False)
        .agg(
            Orders=("OrderID", "count"),
            m2=("M2", "sum"),
            OrderIDs=("OrderID", lambda s: ", ".join(map(str, pd.Series(s).astype(str).tolist()))),
            ReferentieKlant=("ReferentieKlant", lambda s: ", ".join(pd.Series(s).dropna().astype(str).unique().tolist()[:6])) if "ReferentieKlant" in herpland.columns else ("OrderID", lambda s: ""),
        )
        .reset_index()
        .sort_values(["Kleur", "OudePlanDatum", "NieuwePlanDatum", "Verplaatsrichting"], ascending=[True, True, True, True], na_position="last", kind="mergesort")
    )
    verplaatsblokken_kleur["Actieblok"] = (
        "Kleur " + verplaatsblokken_kleur["Kleur"].astype(str)
        + ": " + verplaatsblokken_kleur["Orders"].astype(int).astype(str)
        + " order(s) van " + verplaatsblokken_kleur["OudePlanDatum"].dt.strftime("%d-%m-%Y")
        + " naar " + verplaatsblokken_kleur["NieuwePlanDatum"].dt.strftime("%d-%m-%Y")
    )

actie_cols = [
    "Actie", "Verplaatsgroep", "OudePlanDatum", "NieuwePlanDatum", "Verschil_dagen", "Verplaatsrichting",
    "Naam", "OrderID", "VolgNr", "Kleur", "M2", "LeverDatum", "LaatsteToegestaneDag",
    "VroegsteDag", "Binnen", "Regel_AanleverdatumVanToepassing", "ReferentieKlant", "Omschrijving"
]
herpland_actie = herpland.copy()
actie_cols = [c for c in actie_cols if c in herpland_actie.columns]
herpland_actie = herpland_actie[actie_cols]
if len(herpland_actie):
    herpland_actie = herpland_actie.sort_values([c for c in ["OudePlanDatum", "Kleur", "NieuwePlanDatum", "OrderID"] if c in herpland_actie.columns]).reset_index(drop=True)
    herpland_actie.insert(0, "Volgorde", range(1, len(herpland_actie) + 1))

# ---------- UI ----------
st.subheader("Samenvatting")
st.markdown("<div class='planner-caption'>Overzicht van de planning, capaciteit en herplan-acties voor de planner.</div>", unsafe_allow_html=True)
c1, c2, c3, c4 = st.columns(4)
c1.metric("Totaal orders", len(df))
c2.metric("Gefixeerd", int(df["Gefixeerd"].sum()))
c3.metric("Herpland", int(len(herpland)))
c4.metric("Niet planbaar + advies", int(len(late)))

if overload_days:
    st.warning("⚠️ Gefixeerde orders overschrijden daglimieten op één of meer dagen.")
    warn_df = pd.DataFrame(overload_days, columns=["Dag", "m2_gefixeerd", "kleuren_gefixeerd", "Cap_m2", "Cap_kleuren"])
    warn_df["m2_gefixeerd"] = warn_df["m2_gefixeerd"].apply(fmt_m2)
    warn_df["Cap_m2"] = warn_df["Cap_m2"].apply(fmt_m2)
    render_df(warn_df, height=220)

with st.expander("Toelichting planningsregels"):
    st.write(f"- **Hard deadline**: plandatum ≤ LeverDatum - {int(min_dagen_voor_lever)} dag(en)")
    st.write(f"- **Voorkeur buffer**: optimizer probeert plandatum ≤ LeverDatum - {int(pref_dagen_voor_lever)} dag(en)")
    st.write(
        "- **Nieuwe regel**: als **Binnen = waar**, dan is een order direct planbaar. Als **Binnen = onwaar**, dan is de vroegste plandatum **AanleverDatum + 1 dag**."
    )
    st.write("- **Binnen = waar**: order is direct planbaar vanaf de start van de planningshorizon.")
    st.write("- **AanleverDatum** wordt gebruikt voor de vroegste plandatum als **Binnen = onwaar**: **AanleverDatum + 1 dag**.")

resultaat_tab, verplaatskleur_tab, herplan_tab, detail_tab, niet_planbaar_tab, instellingen_tab = st.tabs(
    ["Resultaat", "Verplaatsblokken op kleur", "Te herplannen orders", "Detailplanning", "Niet planbaar + advies", "Instellingen"]
)

with resultaat_tab:
    st.markdown("### Dagoverzicht")
    dagsamenvatting_view = fmt_df_m2(dagsamenvatting.copy(), ["m2"])
    dagsamenvatting_view["Belasting_m2_pct"] = dagsamenvatting["Belasting_m2_pct"].apply(format_pct)
    dagsamenvatting_view["Belasting_kleur_pct"] = dagsamenvatting["Belasting_kleur_pct"].apply(format_pct)
    dagsamenvatting_view = compact_dagoverzicht_columns(dagsamenvatting_view)
    render_df(dagsamenvatting_view, height=420)

    st.markdown("### Kleurblokken per dag")
    kleurblokken_view = drop_deelorder_columns(fmt_df_m2(kleurblokken.copy(), ["m2"]))
    render_df(kleurblokken_view, height=420)

    st.markdown("### Kleurbelasting per week")
    st.caption("Overzicht van kleurbelasting per weeknummer, zodat de planner snel ziet welke kleuren per week de grootste blokken vormen.")
    if len(kleurbelasting_week) == 0:
        st.info("Geen kleurbelasting per week beschikbaar.")
    else:
        kbw_view = fmt_df_m2(kleurbelasting_week.copy(), ["m2"])
        kbw_cols = ["Week", "WeekNr", "Kleur", "Orders", "m2", "DagenGepland", "EersteDag", "LaatsteDag"]
        kbw_cols = [c for c in kbw_cols if c in kbw_view.columns]
        render_df(kbw_view[kbw_cols], height=420)

with verplaatskleur_tab:
    st.markdown("### Verplaatsblokken op kleur")
    st.caption("Sortering: eerst kleur en daarbinnen op oude plandatum en daarna op nieuwe plandatum. Zo kan de planner eerst per kleur werken.")
    if len(herpland) == 0:
        st.success("Geen orders herpland.")
    else:
        vbk_view = fmt_df_m2(verplaatsblokken_kleur.copy(), ["m2"])
        k1, k2, k3 = st.columns(3)
        k1.metric("Kleurblokken", len(vbk_view))
        k2.metric("Te verplaatsen orders", int(vbk_view["Orders"].sum()) if "Orders" in vbk_view.columns else 0)
        k3.metric("Te verplaatsen m²", fmt_m2(float(vbk_view["m2"].astype(str).str.replace(",", ".", regex=False).astype(float).sum())) if "m2" in vbk_view.columns and len(vbk_view) else "0,0")
        vbk_cols = ["Kleur", "OudePlanDatum", "NieuwePlanDatum", "OrderIDs", "Verplaatsrichting", "Actieblok", "Orders", "m2", "ReferentieKlant"]
        vbk_cols = [c for c in vbk_cols if c in vbk_view.columns]
        render_df(vbk_view[vbk_cols], height=420)

with herplan_tab:
    st.markdown("### Nieuwe opzet voor de planner")
    st.write(
        "Dit tabblad is opgezet voor de werkwijze waarbij orders **één voor één** worden verschoven. "
        "Werk eerst per **oude plandatum** en daarna binnen die dag op **kleur**. Loop daarna de **order-actielijst** in die volgorde door."
    )
    st.markdown(
        f"""
        <div class='planner-card'>
        <b>Praktische werkwijze voor de planner</b><br>
        1. Sorteer op <b>oude plandatum</b>.<br>
        2. Werk daarbinnen op <b>kleur</b> en daarna per <b>oude dag → nieuwe dag</b> alle orders af.<br>
        3. Gebruik daarna de actielijst als afvinklijst voor handmatige verplaatsing in het planningssysteem.
        </div>
        """,
        unsafe_allow_html=True,
    )

    if len(herpland) == 0:
        st.success("Geen orders herpland.")
    else:
        st.markdown("#### 1) Verplaatsblokken (oude plandatum leidend, daarna kleur en doeldag)")
        vb_view = fmt_df_m2(verplaatsblokken_kleur.copy(), ["m2"])
        h1, h2, h3 = st.columns(3)
        h1.metric("Verplaatsblokken", len(vb_view))
        h2.metric("Te verplaatsen orders", int(vb_view["Orders"].sum()) if "Orders" in vb_view.columns else 0)
        h3.metric("Te verplaatsen m²", fmt_m2(vb_view["m2"].astype(str).str.replace(",", ".", regex=False).astype(float).sum()) if "m2" in vb_view.columns and len(vb_view) else "0,0")
        vb_cols = [
            "Kleur", "OudePlanDatum", "NieuwePlanDatum", "OrderIDs", "Verplaatsrichting", "Actieblok", "Orders", "m2", "ReferentieKlant"
        ]
        vb_cols = [c for c in vb_cols if c in vb_view.columns]
        render_df(vb_view[vb_cols], height=320)

        st.markdown("#### 2) Order-actielijst (volgorde voor handmatige verwerking)")
        hap_view = drop_deelorder_columns(fmt_df_m2(herpland_actie.copy(), ["M2"]))
        st.caption("Aanpak: werk eerst per kleurblok en verwerk daarna de order-acties van boven naar beneden.")
        render_df(hap_view, height=520)

with detail_tab:
    st.markdown("### Detailplanning (alles)")
    detail = compact_detailplanning_columns(drop_deelorder_columns(fmt_df_m2(df.copy(), ["M2"])))
    render_df(detail, height=560)

with niet_planbaar_tab:
    st.markdown("### Niet planbaar + advies")
    st.caption("Compact overzicht per adviesblok, vergelijkbaar met verplaatsblokken. Gebruik de detailtabel alleen voor uitzonderingen of orderniveau-controle.")
    if len(late) == 0:
        st.success("Geen te late of niet-planbare orders.")
    else:
        late_full = late.copy()
        adviesblokken = pd.DataFrame()
        grp_cols = [c for c in ["AdviesPlanDatum", "AdviesLeverDatum", "AdviesActie", "Reden"] if c in late_full.columns]
        if len(grp_cols):
            adviesblokken = (
                late_full.groupby(grp_cols, dropna=False)
                .agg(
                    Orders=("OrderID", "count"),
                    Kleuren=("Kleur", pd.Series.nunique),
                    m2=("M2", "sum"),
                    OrderIDs=("OrderID", lambda s: ", ".join(map(str, pd.Series(s).astype(str).tolist()))),
                )
                .reset_index()
                .sort_values([c for c in ["AdviesPlanDatum", "AdviesLeverDatum", "AdviesActie"] if c in grp_cols])
            )
            if "AdviesPlanDatum" in adviesblokken.columns and "AdviesLeverDatum" in adviesblokken.columns:
                adviesblokken["Adviesblok"] = (
                    "Plan " + adviesblokken["AdviesPlanDatum"].dt.strftime("%d-%m-%Y").fillna("n.v.t.")
                    + " | uitlevering " + adviesblokken["AdviesLeverDatum"].dt.strftime("%d-%m-%Y").fillna("n.v.t.")
                )
            elif "AdviesPlanDatum" in adviesblokken.columns:
                adviesblokken["Adviesblok"] = "Plan " + adviesblokken["AdviesPlanDatum"].dt.strftime("%d-%m-%Y").fillna("n.v.t.")
            else:
                adviesblokken["Adviesblok"] = adviesblokken.index.astype(str)

        b1, b2, b3 = st.columns(3)
        b1.metric("Adviesblokken", len(adviesblokken) if len(adviesblokken) else 0)
        b2.metric("Niet planbare orders", int(len(late_full)))
        b3.metric("Niet planbaar m²", fmt_m2(float(late_full["M2"].sum())) if "M2" in late_full.columns and len(late_full) else "0,0")

        if len(adviesblokken):
            st.markdown("#### Compact adviesoverzicht")
            adviesblokken_view = adviesblokken.copy()
            if "LeverDatum" in late_full.columns:
                lever_range = (
                    late_full.groupby(grp_cols, dropna=False)["LeverDatum"]
                    .agg(["min", "max"])
                    .reset_index()
                    .rename(columns={"min": "EersteGevraagdeLeverdatum", "max": "LaatsteGevraagdeLeverdatum"})
                )
                adviesblokken_view = adviesblokken_view.merge(lever_range, on=grp_cols, how="left")
            adviesblokken_view = fmt_df_m2(adviesblokken_view, ["m2"])
            ab_cols = [
                "Adviesblok", "AdviesActie", "Reden", "Orders", "Kleuren", "m2",
                "EersteGevraagdeLeverdatum", "LaatsteGevraagdeLeverdatum"
            ]
            ab_cols = [c for c in ab_cols if c in adviesblokken_view.columns]
            render_df(adviesblokken_view[ab_cols], height=340)

        with st.expander("Orderdetails tonen (alleen bij uitzondering)", expanded=False):
            late_view = drop_deelorder_columns(compact_late_columns(late_full.copy()))
            late_view = fmt_df_m2(late_view, ["M2"]).sort_values([c for c in ["OudePlanDatum", "AdviesPlanDatum", "LeverDatum", "OrderID"] if c in late_view.columns])
            render_df(late_view, height=420)

with instellingen_tab:
    st.markdown("### Gebruikte instellingen")
    instellingen = pd.DataFrame(
        [
            ("Max m² per dag", max_m2),
            ("Max kleuren per dag", max_kleuren),
            ("Vrijdag override actief", "Ja" if use_friday_override else "Nee"),
            ("Max m² op vrijdag", friday_max_m2),
            ("Max kleuren op vrijdag", friday_max_kleuren),
            ("Hard buffer vóór leverdatum", min_dagen_voor_lever),
            ("Voorkeur buffer vóór leverdatum", pref_dagen_voor_lever),
            ("Vroegste dag-regel (indien Binnen = onwaar)", "AanleverDatum + 1 dag"),
            ("Fixeer volgende werkdag", "Ja" if fixeer_morgen else "Nee"),
            ("Fixeer 2e volgende werkdag", "Ja" if fixeer_overmorgen else "Nee"),
            ("Weekenden uitsluiten", "Ja" if exclude_weekends else "Nee"),
            ("Feestdagen/sluitingsdagen uitsluiten", "Ja" if exclude_holidays else "Nee"),
            ("Optimizer", optimizer),
        ("App versie", APP_VERSION),
            ("Aantal ingestelde feestdagen/sluitingsdagen", len(holidays)),
            ("Persistente instellingenbestand", str(SETTINGS_FILE.name)),
        ],
        columns=["Instelling", "Waarde"],
    )
    render_df(instellingen)

    feestdagen_df = pd.DataFrame(sorted(holidays), columns=["Feestdag_of_sluitingsdag"])
    st.markdown("#### Actieve feestdagen / sluitingsdagen")
    render_df(feestdagen_df, height=260)

    if locked_days:
        locked_df = pd.DataFrame(sorted(pd.to_datetime(list(locked_days))), columns=["Gefixeerde_dag"])
        st.markdown("#### Dagen die op slot staan door fixatie")
        render_df(locked_df, height=220)

# ---------- Export ----------
output = BytesIO()
export_detail = compact_detailplanning_columns(drop_deelorder_columns(normalize_dates(df.copy())))
export_herpland = drop_deelorder_columns(normalize_dates(herpland_actie.copy()))
export_verplaatsblokken = drop_deelorder_columns(normalize_dates(verplaatsblokken.copy()))
export_verplaatsblokken_kleur = normalize_dates(verplaatsblokken_kleur.copy())
evbk_cols = ["Kleur", "OudePlanDatum", "NieuwePlanDatum", "OrderIDs", "Verplaatsrichting", "Actieblok", "Orders", "m2", "ReferentieKlant"]
evbk_cols = [c for c in evbk_cols if c in export_verplaatsblokken_kleur.columns]
export_verplaatsblokken_kleur = export_verplaatsblokken_kleur[evbk_cols]
export_dag = compact_dagoverzicht_columns(normalize_dates(dagsamenvatting.copy()))
export_kleur = drop_deelorder_columns(normalize_dates(kleurblokken.copy()))
export_late = drop_deelorder_columns(compact_late_columns(normalize_dates(late.copy())))


export_settings = pd.DataFrame(
    [
        ("Max m² per dag", max_m2),
        ("Max kleuren per dag", max_kleuren),
        ("Vrijdag override actief", "Ja" if use_friday_override else "Nee"),
        ("Max m² op vrijdag", friday_max_m2),
        ("Max kleuren op vrijdag", friday_max_kleuren),
        ("Hard buffer vóór leverdatum", min_dagen_voor_lever),
        ("Voorkeur buffer vóór leverdatum", pref_dagen_voor_lever),
        ("Vroegste dag-regel (indien Binnen = onwaar)", "AanleverDatum + 1 dag"),
        ("Fixeer volgende werkdag", "Ja" if fixeer_morgen else "Nee"),
        ("Fixeer 2e volgende werkdag", "Ja" if fixeer_overmorgen else "Nee"),
        ("Weekenden uitsluiten", "Ja" if exclude_weekends else "Nee"),
        ("Feestdagen/sluitingsdagen uitsluiten", "Ja" if exclude_holidays else "Nee"),
        ("Optimizer", optimizer),
        ("App versie", APP_VERSION),
        ("Aantal ingestelde feestdagen/sluitingsdagen", len(holidays)),
        ("Persistente instellingenbestand", str(SETTINGS_FILE.name)),
        ("Feestdagen/sluitingsdagen (tekstveld)", st.session_state["holiday_text"]),
    ],
    columns=["Instelling", "Waarde"],
)

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    export_detail.to_excel(writer, index=False, sheet_name="Resultaat")
    export_herpland.to_excel(writer, index=False, sheet_name="Te herplannen orders")
    export_verplaatsblokken.to_excel(writer, index=False, sheet_name="Verplaatsblokken")
    export_verplaatsblokken_kleur.to_excel(writer, index=False, sheet_name="Verplaatsblokken kleur")
    export_dag.to_excel(writer, index=False, sheet_name="Dagoverzicht")
    export_kleur.to_excel(writer, index=False, sheet_name="Kleurblokken")
    export_late.to_excel(writer, index=False, sheet_name="Niet planbaar advies")
    export_settings.to_excel(writer, index=False, sheet_name="Instellingen")
    if overload_days:
        pd.DataFrame(overload_days, columns=["Dag", "m2_gefixeerd", "kleuren_gefixeerd", "Cap_m2", "Cap_kleuren"]).to_excel(
            writer, index=False, sheet_name="Fix-overload"
        )

    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_border = Border(bottom=thin)
    date_format = "DD-MM-YYYY"
    m2_format = "0.0"
    pct_format = "0%"
    center_headers = {"Stoplicht_m2", "Stoplicht_kleuren", "Gefixeerd", "Gefixeerd_JaNee", "Regel_AanleverdatumVanToepassing"}
    wide_caps = {"ReferentieKlant": 24, "Omschrijving": 24, "OrderIDs": 34, "Waarde": 60, "Actie": 22, "Actieblok": 28, "Reden": 22, "Klanten": 24, "AdviesToelichting": 34}

    for wsname in wb.sheetnames:
        ws = wb[wsname]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = data_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for j, header in enumerate(headers, start=1):
            if header is None:
                continue
            h = str(header)
            h_low = h.lower()
            for row in ws.iter_rows(min_row=2, min_col=j, max_col=j):
                cell = row[0]
                if cell.value is None:
                    continue
                if ("datum" in h_low) or h_low.endswith("dag") or h_low.startswith("feestdag") or h_low.startswith("gefixeerde_dag"):
                    if hasattr(cell.value, "year"):
                        cell.number_format = date_format
                if ("m2" in h_low) or ("m²" in h_low):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = m2_format
                if "belasting_" in h_low and h_low.endswith("_pct"):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = pct_format
                if h in center_headers:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        for col_cells in ws.columns:
            col_idx = col_cells[0].column
            col_letter = get_column_letter(col_idx)
            head = str(col_cells[0].value) if col_cells and col_cells[0].value is not None else ""
            max_len = len(head)
            for cell in col_cells[1:]:
                if cell.value is None:
                    continue
                rendered = cell.value.strftime("%d-%m-%Y") if hasattr(cell.value, "strftime") and (("Datum" in head) or head.endswith("Dag")) else str(cell.value)
                max_len = max(max_len, len(rendered))
            width = max(11, min(80, max_len + 2))
            if head in wide_caps:
                width = min(max(width, 18), wide_caps[head])
            if head in ["Stoplicht_m2", "Stoplicht_kleuren"]:
                width = 12
            ws.column_dimensions[col_letter].width = width


output.seek(0)
st.download_button("Download resultaat (Excel)", data=output.getvalue(), file_name="planning_resultaat.xlsx")
